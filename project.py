import sys

import sqlite3
import csv
import xlsxwriter
from openpyxl import load_workbook
from PIL import Image, ImageFont, ImageDraw

from os import path, rename, mkdir, remove
from shutil import rmtree
from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QMessageBox, QTableWidgetItem, QFileDialog
from PyQt6.QtCore import Qt
from PyQt6 import uic
filtering_history, index = [''], 0


class AddCertificate(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi("ui_files/add_window.ui", self)

        self.widget_main = None
        self.main_window = None
        self.button_add.clicked.connect(self.add_certificate)
        self.button_select_file.clicked.connect(self.add_certificate_from_file)

    def add_certificate(self):
        con = sqlite3.connect('certificate_db.sqlite')
        cur = con.cursor()
        data_list = list(filter(lambda txt: txt != '', self.data_input.toPlainText().split('\n')))
        if data_list:
            try:
                if self.check_correct_data(data_list):
                    cur.execute(f"""INSERT INTO certificates VALUES{tuple(data_list)}""").fetchall()
                    con.commit()
                    con.close()
                    valid = QMessageBox.question(
                        self, '', 'Сертификат успешно добавлен.\nХотите продолжить?',
                        QMessageBox.StandardButton.Yes, QMessageBox.StandardButton.No)
                    if valid == QMessageBox.StandardButton.No:
                        self.close_window()
                else:
                    self.message_error()
            except sqlite3.IntegrityError:
                QMessageBox.information(
                    self, '', 'Сертификат с таким номером уже существует.',
                    QMessageBox.StandardButton.Ok)
            except sqlite3.OperationalError:
                self.message_error()
        else:
            self.question_message()
        self.data_input.clear()

    def message_error(self):
        valid = QMessageBox.question(
            self, '', 'Неправильные данные.\nХотите закончить?',
            QMessageBox.StandardButton.Yes, QMessageBox.StandardButton.No)
        if valid == QMessageBox.StandardButton.Yes:
            self.close_window()

    def add_certificate_from_file(self):
        try:
            file = QFileDialog.getOpenFileName(self, '', '')[0]
            if file.split('/')[-1].split('.')[1] == 'csv':
                self.add_csv(file)
            elif file.split('/')[-1].split('.')[1] == 'xlsx':
                self.add_xlsx(file)
            else:
                QMessageBox.information(
                    self, '', 'Выберете файл с расширением csv или xlsx', QMessageBox.StandardButton.Ok)
        except IndexError:
            pass

    def add_csv(self, file):
        with open(file, encoding='utf-8') as csvfile:
            reader = list(filter(lambda t: t, csv.reader(csvfile, delimiter=';', quotechar='"')))
            con = sqlite3.connect('certificate_db.sqlite')
            cur = con.cursor()
            count, flag = 0, True
            if reader:
                for lst in reader:
                    if self.check_correct_data(lst):
                        try:
                            cur.execute(f"""INSERT INTO certificates VALUES{tuple(lst)}""").fetchall()
                        except sqlite3.IntegrityError:
                            count += 1
                        except sqlite3.OperationalError:
                            flag = False
                    else:
                        flag = False
            else:
                self.question_message()
            con.commit()
            con.close()
            self.check_data_entry_valid(count, reader, flag)

    def add_xlsx(self, file):
        con = sqlite3.connect("certificate_db.sqlite")
        cur = con.cursor()
        count, flag = 0, True

        wookbook = load_workbook(file)
        worksheet = wookbook.active
        table_data = []

        for i in range(0, worksheet.max_row):
            row = []
            for col in worksheet.iter_cols(1, worksheet.max_column):
                row.append(col[i].value)
            table_data.append(row)
        table_data = list(filter(lambda l: l[0] is not None, table_data))
        if table_data:
            for lst in table_data:
                if self.check_correct_data(lst):
                    try:
                        cur.execute(f"""INSERT INTO certificates VALUES{tuple(lst)}""").fetchall()
                    except sqlite3.IntegrityError:
                        count += 1
                    except sqlite3.OperationalError:
                        flag = False
                else:
                    flag = False
        else:
            self.question_message()
        con.commit()
        con.close()
        self.check_data_entry_valid(count, table_data, flag)

    def check_data_entry_valid(self, count, data, flag):
        if count == len(data):
            if count == 1:
                QMessageBox.information(
                    self, '', 'Сертификат с таким номером уже существует.',
                    QMessageBox.StandardButton.Ok)
            else:
                QMessageBox.information(
                    self, '', 'Сертификаты с такими номерами уже существуют.',
                    QMessageBox.StandardButton.Ok)
        elif not flag:
            valid = QMessageBox.question(
                self, '', 'Неправильные данные.\nХотите закончить?',
                QMessageBox.StandardButton.Yes, QMessageBox.StandardButton.No)
            if valid == QMessageBox.StandardButton.Yes:
                self.close_window()
        else:
            self.success_add()

    def close_window(self):
        self.data_input.clear()
        self.close()
        self.widget_main = CertificateNavigation()
        self.widget_main.show()

    def question_message(self):
        valid = QMessageBox.question(
            self, '', 'Вы ничего не вписали.\nХотите вернуться?',
            QMessageBox.StandardButton.Yes, QMessageBox.StandardButton.No)
        if valid == QMessageBox.StandardButton.Yes:
            self.close_window()

    def success_add(self):
        valid = QMessageBox.question(
            self, '', 'Данные успешно добавлены.\nХотите продолжить?',
            QMessageBox.StandardButton.Yes, QMessageBox.StandardButton.No)
        if valid == QMessageBox.StandardButton.No:
            self.close_window()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Escape:
            self.main_window = CertificateNavigation()
            self.main_window.show()
            self.close()

    def check_correct_data(self, data):
        try:
            for i in range(len(data)):
                if i == 0 or i == 6 or i == 7:
                    data[i] = int(data[i])
            ru_alp = "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"
            rating = ['отлично', 'хорошо', 'удовлетворительно']
            string_data = list(filter(lambda elem: elem == str(elem), data))
            for i in range(0, 3):
                if not (string_data[i] == string_data[i].capitalize() and set(string_data[i].lower()) <= set(ru_alp)):
                    return False
            if string_data[4] not in rating:
                return False
            for i in range(5, 7):
                if not (string_data[i] == string_data[i].capitalize() and set(string_data[i].lower()) <= set(ru_alp)):
                    return False
            return True
        except ValueError:
            return False


class ViewCertificate(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi("ui_files/view_window.ui", self)

        global filtering_history
        self.main_window = None
        self.result = None
        self.key_event = False
        self.flag_first_filter = True

        self.con = sqlite3.connect('certificate_db.sqlite')
        self.cur = self.con.cursor()

        self.button_find.clicked.connect(self.find_certificate)
        self.button_convert.clicked.connect(self.open_main_window)
        self.button_convert.hide()

    def find_certificate(self):
        data_list = list(filter(lambda txt: txt != '', self.data_input.toPlainText().split('\n')))
        try:
            if data_list:
                self.result = self.cur.execute(f"""SELECT * FROM certificates WHERE {' AND '.join(data_list)}
                                                """).fetchall()
                if self.data_input.toPlainText() != filtering_history[-1]:
                    filtering_history.append(self.data_input.toPlainText())
                    self.flag_first_filter = True
            else:
                self.result = self.cur.execute(f"""SELECT * FROM certificates""").fetchall()
            self.data_output.setRowCount(len(self.result))
            if self.result:
                self.data_output.setColumnCount(len(self.result[0]))
                for i, elem in enumerate(self.result):
                    for j, val in enumerate(elem):
                        self.data_output.setItem(i, j, QTableWidgetItem(str(val)))
                self.button_convert.show()
            else:
                self.clear_data()
                self.question_message()
        except sqlite3.OperationalError:
            valid = QMessageBox.question(
                self, '', 'Неправильные данные.\nХотите вернуться?',
                QMessageBox.StandardButton.Yes, QMessageBox.StandardButton.No)
            self.button_convert.hide()
            self.clear_data()
            if valid == QMessageBox.StandardButton.Yes:
                self.close_window()

    def close_window(self):
        self.clear_data()
        self.close()
        self.main_window = CertificateNavigation()
        self.main_window.show()

    def clear_data(self):
        self.data_input.clear()
        self.data_output.clear()
        self.data_output.setRowCount(0)
        self.data_output.setColumnCount(0)

    def question_message(self):
        if not self.key_event:
            self.button_convert.hide()
            valid = QMessageBox.question(
                self, '', 'Ничего не нашлось.\nХотите вернуться?',
                QMessageBox.StandardButton.Yes, QMessageBox.StandardButton.No)
            if valid == QMessageBox.StandardButton.Yes:
                self.close_window()

    def keyPressEvent(self, event):
        global index
        if event.key() == Qt.Key.Key_Escape:
            self.main_window = CertificateNavigation()
            self.main_window.show()
            self.close()
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            if filtering_history:
                flag = True
                if self.flag_first_filter:
                    if self.data_input.toPlainText() in filtering_history:
                        index = len(filtering_history) - \
                                filtering_history[::-1].index(self.data_input.toPlainText()) - 1
                        flag = True
                    else:
                        flag = False
                    self.flag_first_filter = False
                if event.key() == Qt.Key.Key_U:
                    if not flag or (filtering_history[index] != self.data_input.toPlainText()):
                        index = 0
                    if index == 0:
                        index = len(filtering_history) - 1
                    else:
                        index -= 1
                    self.data_input.setPlainText(filtering_history[index])
                elif event.key() == Qt.Key.Key_D:
                    if not flag or (filtering_history[index] != self.data_input.toPlainText()):
                        index = len(filtering_history) - 1
                    if index == len(filtering_history) - 1:
                        index = 0
                    else:
                        index += 1
                    self.data_input.setPlainText(filtering_history[index])
        if self.result:
            if event.key() == Qt.Key.Key_Backspace:
                rows = list(set([i.row() for i in self.data_output.selectedItems()]))
                if rows:
                    ids = [self.data_output.item(i, 0).text() for i in rows]
                    if len(ids) == 1:
                        self.cur.execute(f"""DELETE FROM certificates WHERE id={ids[0]}""").fetchall()
                    else:
                        self.cur.execute(f"""DELETE FROM certificates WHERE id IN {tuple(ids)}""").fetchall()
                    self.con.commit()
                    self.key_event = True
                    self.button_convert.hide()
                    self.find_certificate()
                    QMessageBox.information(self, '', 'Сертификат успешно удален.',
                                            QMessageBox.StandardButton.Ok)
                    self.key_event = False

    def open_main_window(self):
        rows = list(set([i.row() for i in self.data_output.selectedItems()]))
        if rows:
            ids = [self.data_output.item(i, 0).text() for i in rows]
            self.clear_data()
            self.main_window = CertificateNavigation(ids)
            self.main_window.show()
            self.close()


class ConvertCertificate(QWidget):
    def __init__(self, ids=None):
        super().__init__()
        uic.loadUi("ui_files/convert_window.ui", self)

        self.main_window = None
        self.ids = ids
        self.result = None

        self.button_convert_csv.clicked.connect(self.convert_csv)
        self.button_convert_xlsx.clicked.connect(self.convert_xlsx)
        self.button_convert_jpg.clicked.connect(self.convert_jpg)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Escape:
            self.main_window = CertificateNavigation(self.ids)
            self.main_window.show()
            self.close()

    def connect_bd(self):
        con = sqlite3.connect("certificate_db.sqlite")
        cur = con.cursor()
        if len(self.ids) == 1:
            self.result = cur.execute(f"""SELECT * FROM certificates WHERE id={self.ids[0]}""").fetchall()
        else:
            self.result = cur.execute(f"""SELECT * FROM certificates WHERE id IN{tuple(self.ids)}""").fetchall()
        con.close()

    def save_to_directory(self, name):
        try:
            if path.isfile(name):
                QMessageBox.information(
                    self, "", f"{name.split('.')[1]}-файл успешно создан.\nВыберете куда хотите его сохранить.",
                    QMessageBox.StandardButton.Ok)
                directory = QFileDialog.getExistingDirectory(self, '', '') + f"/{name}"
                rename(name, directory)
                QMessageBox.information(
                    self, "", f"Файл успешно сохранен.",
                    QMessageBox.StandardButton.Ok)
            else:
                QMessageBox.information(
                    self, "", f"Папка успешно создана.\nВыберете куда хотите её сохранить.",
                    QMessageBox.StandardButton.Ok)
                directory = QFileDialog.getExistingDirectory(self, '', '') + f"/{name}"
                if path.exists(directory):
                    rmtree(directory)
                rename(name, directory)
                QMessageBox.information(
                    self, "", f"Папка успешно сохранена.",
                    QMessageBox.StandardButton.Ok)
        except OSError:
            if path.isfile(name):
                remove(name)
            else:
                rmtree(name)

    def convert_csv(self):
        self.connect_bd()
        with open("certificate_db.csv", 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(
                csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            for pers in self.result:
                writer.writerow(list(pers))
            self.save_to_directory("certificate_db.csv")

    def convert_xlsx(self):
        self.connect_bd()
        workbook = xlsxwriter.Workbook("certificate_db.xlsx")
        worksheet = workbook.add_worksheet()
        for row, pers in enumerate(self.result):
            for i in range(len(pers)):
                worksheet.write(row, i, pers[i])
        workbook.close()
        self.save_to_directory("certificate_db.xlsx")

    def convert_jpg(self):
        self.connect_bd()
        length = len(self.result)
        if not path.exists("Сертификаты") and length > 1:
            mkdir("Сертификаты")
        for i in range(length):
            name, patronymic, surname = self.result[i][1], self.result[i][2], self.result[i][3]
            im = Image.open(f'templates/{self.result[i][5]}.jpg')
            font1 = ImageFont.truetype('/System/Library/Fonts/Helvetica.ttc', size=80)
            font2 = ImageFont.truetype('/System/Library/Fonts/Helvetica.ttc', size=35)
            font3 = ImageFont.truetype('/System/Library/Fonts/Helvetica.ttc', size=30)
            font4 = ImageFont.truetype('/System/Library/Fonts/Helvetica.ttc', size=40)
            draw = ImageDraw.Draw(im)
            draw.text((255, 1305), f'{name} {patronymic} {surname}', font=font1, fill='#272026')
            draw.text((580, 1559), f'4 октября {self.result[i][6]} года - 10 мая {self.result[i][7]} года.',
                      font=font2, fill='#272026')
            draw.text((770, 1648), f'{self.result[i][0]}', font=font3, fill='#272026')
            draw.text((1248, 1402), f"«{self.result[i][4].split()[0]}", font=font4, fill='#272026')
            draw.text((250, 1445), f"{' '.join(self.result[i][4].split()[1:])}» в рамках проекта",
                      font=font4, fill='#272026')
            draw.text((250, 1495), "«Лицей Академии Яндекса».", font=font4, fill="#272026")

            if length > 1:
                im.save(f"Сертификаты/Сертификат_{surname}_{name}_{patronymic}.jpg")
            else:
                im.save(f"Сертификат_{surname}_{name}_{patronymic}.jpg")
                self.save_to_directory(f"Сертификат_{surname}_{name}_{patronymic}.jpg")
        if length > 1:
            self.save_to_directory("Сертификаты")


class CertificateNavigation(QMainWindow):
    def __init__(self, ids=None):
        super().__init__()
        uic.loadUi("ui_files/main_window.ui", self)

        self.widget_add = None
        self.widget_view = None
        self.widget_convert = None

        self.add_certificate.clicked.connect(self.open_add_window)
        self.view_certificate.clicked.connect(self.open_view_window)
        self.button_convert.clicked.connect(self.open_convert_window)
        self.button_convert.move(190, 90)
        self.ids = ids
        if self.ids:
            self.button_convert.show()
        else:
            self.button_convert.hide()

    def open_add_window(self):
        self.widget_add = AddCertificate()
        self.widget_view = ViewCertificate()
        self.widget_view.close()
        self.widget_add.show()
        self.close()

    def open_view_window(self):
        self.widget_view = ViewCertificate()
        self.widget_add = AddCertificate()
        self.widget_add.close()
        self.button_convert.hide()
        self.close()
        self.widget_view.show()

    def open_convert_window(self):
        self.widget_add = AddCertificate()
        self.widget_view = ViewCertificate()
        self.widget_add.close()
        self.widget_view.close()
        self.widget_convert = ConvertCertificate(self.ids)
        self.widget_convert.show()
        self.close()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    nav = CertificateNavigation()
    nav.show()
    sys.exit(app.exec())
